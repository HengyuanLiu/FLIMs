from openpyxl import load_workbook
import csv
import os
def save_matrix_to_csv(matrix, filename):
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["mutant_id"] + [f"{i+1}" for i in range(5)])
        writer.writerows(matrix)
def get_top50_xlsx(file_path):
   
    results = []
    
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # 跳过标题行（假设首行为标题）
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 6:
                continue
                
            try:
                # 处理第六列（索引5）的值
                key = float(row[5])
                results.append( (key, row[0]) )
            except (TypeError, ValueError):
                continue
                
        # 排序逻辑：主列降序 + 次列升序
        sorted_results = sorted(results, key=lambda x: (-x[0], str(x[1])))
        return [item[1] for item in sorted_results[:50]]
    
    except FileNotFoundError:
        print(f"error {file_path} not exist")
   
    except Exception as e:
        print(f"error：{str(e)}")
    finally:
        wb.close() if 'wb' in locals() else None
    
    return []

def txt_to_list(file_path):
    
    result = []
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read().strip()  # 读取并去除首尾空白
            
            # 处理多种分隔符情况（逗号+空格、制表符、多个空格等）
            elements = [s.strip() for s in content.replace('\t', ' ').split(',')]
            
            # 转换为整数并过滤无效项
            for num_str in elements:
                if num_str:  # 跳过空字符串
                    try:
                        result.append(int(num_str))
                    except ValueError:
                        print(f"warning '{num_str}' skip")
                        
    except FileNotFoundError:
        print(f"{file_path} not exist")
    except Exception as e:
        print(f"error {str(e)}")
    
    return result

if __name__ == "__main__":
    buglist_dict = {
      'Chart': {'total': 1, 'start': 1},
    #   'Math': {'total': 107, 'start': 1},
    #  'Mockito': {'total': 38, 'start': 1},
    # 'Time': {'total': 26, 'start': 1},
    #    'Lang': {'total': 65, 'start': 1},
    #  'Closure': {'total': 134, 'start': 1},
    }
    # modelslist=['deepseek-14b']
    # modelslist=['deepseek-7b','deepseek-14b','deepseek-llama-8b','llama-8b','qwen-7b-nowt','qwen-14b','qwen-coder-7b','qwen-coder-14b']
    # modelslist=['deepseek-7b','deepseek-14b','deepseek-llama-8b','qwen-7b','qwen-14b','qwen-coder-7b','qwen-coder-14b']
    # for models in modelslist:
    for bug, info in buglist_dict.items():
            total = info['total']
            start = info['start']
            
            for i in range(start, 1 + total):
                print(f"{bug}_{i}")
                mutant_id_path=rf"D:\Graduate\code\result\error1\{bug}\{i}\pf_{i}.xlsx"
                wkill_mutant_id1=rf"D:\Graduate\new-experimina\result\deepseek-8b\id-static\output5\{bug}\{bug}_{i}.txt"
                wkill_mutant_id2=rf"D:\Graduate\new-experimina\result\deepseek-8b\id-static\output1\{bug}\{bug}_{i}.txt"
                wkill_mutant_id3=rf"D:\Graduate\new-experimina\result\deepseek-8b\id-static\output2\{bug}\{bug}_{i}.txt"
                wkill_mutant_id4=rf"D:\Graduate\new-experimina\result\deepseek-8b\id-static\output3\{bug}\{bug}_{i}.txt"
                wkill_mutant_id5=rf"D:\Graduate\new-experimina\result\deepseek-8b\id-static\output4\{bug}\{bug}_{i}.txt"
                outpath=rf"D:\Graduate\new-experimina\result\deepseek-8b\wkill_matrix\{bug}"
                if not os.path.exists(outpath):
                    os.makedirs(outpath)
                output_file = rf"{outpath}\{bug}_{i}_wkill_matrix.csv"
                result = get_top50_xlsx(mutant_id_path)
                sample_list1 = txt_to_list(wkill_mutant_id1)
                sample_list2 = txt_to_list(wkill_mutant_id2)
                sample_list3 = txt_to_list(wkill_mutant_id3)
                sample_list4 = txt_to_list(wkill_mutant_id4)
                sample_list5 = txt_to_list(wkill_mutant_id5)
                
                sets = [lst for lst in [sample_list1, sample_list2, sample_list3, sample_list4, sample_list5]]
                # print(sets)
                matrix = []
                for num in result:
                
                    row = [num]  
                    row += [0 if num in s else 1 for s in sets]  
                    matrix.append(row)
                
            
                save_matrix_to_csv(matrix,output_file)